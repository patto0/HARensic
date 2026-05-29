"""
export_excel.py — Forensic Multi-Sheet Excel Workbook Exporter
==============================================================

Produces ONE publication-ready .xlsx workbook containing a separate
worksheet for every artifact section (A_identity … E_urls / E_ai_urls).

Architecture guarantees
-----------------------
- Each sheet is built from its own isolated DataFrame — no merging.
- SECTION_SCHEMAS field whitelists are enforced before any data touches
  the workbook; URL-enrichment fields are physically absent from A–D sheets.
- The exporter calls strip_section_rows() so schema isolation is enforced
  at the record level, not just at column selection time.
- A synthetic "Summary" sheet is prepended with per-section artifact counts,
  attribution breakdowns, and direct hyperlinks to each section tab.

Sheet styling
-------------
- Frozen header row (row 1 is always visible while scrolling).
- Bold, white-text headers on a section-specific accent colour.
- Alternating row fill (white / very-light-grey) for readability.
- Auto-fitted column widths (capped at 80 chars to avoid absurd widths).
- Arial 10 pt body font throughout.
- Summary sheet uses a dark-navy header bar.
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from parsers.category_labels import (
    SECTION_LABELS,
    SECTION_ORDER,
    normalize_results,
    section_label,
)
from parsers.schemas import SECTION_SCHEMAS, strip_section_rows

# ─────────────────────────────────────────────────────────────────────────────
#  COLOUR PALETTE  (one accent per section)
# ─────────────────────────────────────────────────────────────────────────────

# (header_bg_hex, header_font_hex, tab_colour_hex)
_SECTION_COLOURS: Dict[str, Tuple[str, str, str]] = {
    "A_identity":   ("1F4E79", "FFFFFF", "1F4E79"),   # Deep navy
    "B_prompt":     ("375623", "FFFFFF", "375623"),   # Forest green
    "C_security":   ("7B2C2C", "FFFFFF", "7B2C2C"),   # Deep crimson
    "D_autonomous": ("4A3568", "FFFFFF", "4A3568"),   # Deep purple
    "E_urls":       ("0D4A4A", "FFFFFF", "0D4A4A"),   # Deep teal (AI URLs)
}

# Excel tab/sheet display name for each canonical key
_SHEET_DISPLAY_NAMES: Dict[str, str] = {
    "A_identity":   "A_identity",
    "B_prompt":     "B_prompt",
    "C_security":   "C_security",
    "D_autonomous": "D_autonomous",
    "E_urls":       "E. URLs - Domains Visited",  # ≤31 chars; / is illegal in Excel tab names
}
_SUMMARY_COLOURS = ("1A1A2E", "FFFFFF", "1A1A2E")   # Near-black

_ALT_ROW_FILL   = PatternFill("solid", start_color="F5F5F5", fgColor="F5F5F5")
_WHITE_FILL     = PatternFill("solid", start_color="FFFFFF",  fgColor="FFFFFF")
_BODY_FONT      = Font(name="Arial", size=10)
_WRAP_ALIGN     = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGN      = Alignment(vertical="top")

_THIN_SIDE      = Side(style="thin", color="D0D0D0")
_CELL_BORDER    = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)


# ─────────────────────────────────────────────────────────────────────────────
#  INTERNAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _header_style(hex_bg: str, hex_fg: str) -> Tuple[PatternFill, Font, Alignment]:
    fill = PatternFill("solid", start_color=hex_bg, fgColor=hex_bg)
    font = Font(name="Arial", size=10, bold=True, color=hex_fg)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, align


def _auto_col_widths(ws, max_width: int = 80) -> None:
    """Set each column's width to fit its widest cell value (capped)."""
    for col_cells in ws.columns:
        best = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > best:
                best = length
        # header row: always at least the header length
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            min(max(best + 2, 12), max_width)
        )


def _write_section_sheet(
    wb: Workbook,
    section: str,
    rows: List[Dict[str, Any]],
    debug: bool = False,
) -> None:
    """
    Write one artifact section to a dedicated worksheet.

    Steps
    -----
    1. Strip rows to section schema (schema isolation enforced here).
    2. Resolve field list from SECTION_SCHEMAS.
    3. Write header row with accent colour.
    4. Write data rows with alternating fill.
    5. Freeze header, auto-fit columns, set tab colour.
    """
    # ── 1. Schema isolation ──────────────────────────────────────────────
    clean_rows = strip_section_rows(section, rows, debug=debug)
    fields     = SECTION_SCHEMAS.get(section, list(SECTION_SCHEMAS["A_identity"]))

    # ── 2. Create sheet ───────────────────────────────────────────────────
    label    = SECTION_LABELS.get(section, section)
    tab_name = _SHEET_DISPLAY_NAMES.get(section, section)  # human-readable tab

    ws = wb.create_sheet(title=tab_name)

    # Tab colour
    colours = _SECTION_COLOURS.get(section, ("444444", "FFFFFF", "444444"))
    ws.sheet_properties.tabColor = colours[2]

    # ── 3. Header row ─────────────────────────────────────────────────────
    hdr_fill, hdr_font, hdr_align = _header_style(colours[0], colours[1])
    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = _CELL_BORDER

    # ── 4. Data rows ─────────────────────────────────────────────────────
    for row_idx, record in enumerate(clean_rows, start=2):
        fill = _ALT_ROW_FILL if row_idx % 2 == 0 else _WHITE_FILL
        for col_idx, field in enumerate(fields, start=1):
            raw_val = record.get(field, "")
            # Stringify complex values (lists / dicts from chain fields)
            if isinstance(raw_val, (list, dict)):
                raw_val = str(raw_val)
            cell = ws.cell(row=row_idx, column=col_idx, value=raw_val)
            cell.fill      = fill
            cell.font      = _BODY_FONT
            cell.alignment = _WRAP_ALIGN if field in ("reason", "value") else _TOP_ALIGN
            cell.border    = _CELL_BORDER

    # ── 5. Polish ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"           # header always visible
    _auto_col_widths(ws)
    # Row height: allow wrapped text in reason/value columns to breathe
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30


def _write_summary_sheet(
    wb: Workbook,
    results: Dict[str, List[Dict]],
    platform: str,
    har_filename: str,
    cg: int,
    gm: int,
    cl: int,
) -> None:
    """
    Prepend a 'Summary' worksheet with:
    - Run metadata (platform, source file, timestamp)
    - Per-section artifact counts
    - Attribution breakdown (AI / HUMAN / PLATFORM) per section
    - Hyperlinks to each section sheet
    """
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_properties.tabColor = _SUMMARY_COLOURS[2]

    hdr_fill, hdr_font, hdr_align = _header_style(
        _SUMMARY_COLOURS[0], _SUMMARY_COLOURS[1]
    )

    # ── Metadata block ────────────────────────────────────────────────────
    meta_rows = [
        ("HARensic", "Forensic Report"),
        ("Platform",        platform.upper()),
        ("Source File",     har_filename),
        ("Generated",       datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Detection Scores",
         f"ChatGPT={cg}  /  Gemini={gm}  /  Claude={cl}"),
    ]
    title_font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    meta_key_f  = Font(name="Arial", size=10, bold=True,  color="1A1A2E")
    meta_val_f  = Font(name="Arial", size=10, bold=False, color="1A1A2E")

    for r_idx, (key, val) in enumerate(meta_rows, start=1):
        kc = ws.cell(row=r_idx, column=1, value=key)
        vc = ws.cell(row=r_idx, column=2, value=val)
        if r_idx == 1:
            kc.fill = hdr_fill; kc.font = title_font; kc.alignment = hdr_align
            vc.fill = hdr_fill; vc.font = title_font; vc.alignment = hdr_align
            ws.merge_cells(start_row=1, start_column=2,
                           end_row=1, end_column=6)
        else:
            kc.font = meta_key_f
            vc.font = meta_val_f
        kc.border = _CELL_BORDER
        vc.border = _CELL_BORDER

    # ── Section table header ──────────────────────────────────────────────
    tbl_start = len(meta_rows) + 2    # one blank row gap
    tbl_headers = [
        "Section", "Display Name", "Artifacts",
        "AI", "HUMAN", "PLATFORM", "Schema Fields",
    ]
    for c_idx, h in enumerate(tbl_headers, start=1):
        cell = ws.cell(row=tbl_start, column=c_idx, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = _CELL_BORDER

    # ── Section rows ──────────────────────────────────────────────────────
    row_idx = tbl_start + 1
    total_artifacts = 0

    for section in SECTION_ORDER:
        section_rows = results.get(section, [])
        if not section_rows:
            continue

        ai_c  = sum(1 for r in section_rows if r.get("attribution", "").upper() == "AI")
        hu_c  = sum(1 for r in section_rows if r.get("attribution", "").upper() == "HUMAN")
        pl_c  = len(section_rows) - ai_c - hu_c
        n_fld = len(SECTION_SCHEMAS.get(section, []))

        colours = _SECTION_COLOURS.get(section, ("444444", "FFFFFF", "444444"))
        row_fill = PatternFill("solid", start_color=colours[0], fgColor=colours[0])
        row_font = Font(name="Arial", size=10, color=colours[1])
        num_font = Font(name="Arial", size=10, bold=True, color=colours[1])

        values = [
            section,
            SECTION_LABELS.get(section, section),
            len(section_rows),
            ai_c,
            hu_c,
            pl_c,
            n_fld,
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.fill      = row_fill
            cell.font      = num_font if isinstance(val, int) else row_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _CELL_BORDER

        # Hyperlink in the Section column → that sheet
        display = _SHEET_DISPLAY_NAMES.get(section, section)
        ws.cell(row=row_idx, column=1).hyperlink = f"#{display}!A1"
        ws.cell(row=row_idx, column=1).style     = "Hyperlink"
        ws.cell(row=row_idx, column=1).font = Font(
            name="Arial", size=10, color=colours[1],
            bold=True, underline="single",
        )

        total_artifacts += len(section_rows)
        row_idx += 1

    # ── Totals row ────────────────────────────────────────────────────────
    totals_fill = PatternFill("solid", start_color="2C2C2C", fgColor="2C2C2C")
    totals_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    totals_align = Alignment(horizontal="center", vertical="center")

    for c_idx in range(1, len(tbl_headers) + 1):
        cell = ws.cell(row=row_idx, column=c_idx)
        cell.fill      = totals_fill
        cell.font      = totals_font
        cell.alignment = totals_align
        cell.border    = _CELL_BORDER

    ws.cell(row=row_idx, column=1, value="TOTAL").fill   = totals_fill
    ws.cell(row=row_idx, column=1).font = totals_font
    ws.cell(row=row_idx, column=1).alignment = totals_align
    ws.cell(row=row_idx, column=3, value=total_artifacts).fill = totals_fill
    ws.cell(row=row_idx, column=3).font = totals_font
    ws.cell(row=row_idx, column=3).alignment = totals_align

    # ── Polish ────────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[1].height = 28


# ─────────────────────────────────────────────────────────────────────────────
#  PUBLIC API
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(
    results: Dict[str, List[Dict]],
    platform: str,
    har_path: str,
    output_dir: str,
    cg: int = 0,
    gm: int = 0,
    cl: int = 0,
    debug: bool = False,
    filename: Optional[str] = None,
) -> str:
    """
    Export all artifact sections into one multi-sheet Excel workbook.

    Parameters
    ----------
    results    : canonical results dict {"A_identity": [...], ...}
    platform   : detected platform string, e.g. "chatgpt"
    har_path   : source HAR file path (used for naming)
    output_dir : destination directory
    cg/gm/cl   : legacy detection scores for summary sheet
    debug      : propagate to schema validator (prints violations)
    filename   : override the output filename (optional)

    Returns
    -------
    Absolute path to the written .xlsx file.

    Schema isolation
    ----------------
    Each worksheet is built independently via _write_section_sheet(), which
    calls strip_section_rows() before touching the workbook. URL fields are
    physically absent from A–D sheets — not merely hidden or empty-columned.
    """
    os.makedirs(output_dir, exist_ok=True)

    results = normalize_results(results)

    if filename is None:
        base = os.path.splitext(os.path.basename(har_path))[0]
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{base}_{platform}_{ts}.xlsx"

    out_path = os.path.join(output_dir, filename)

    wb = Workbook()
    # Remove the default empty sheet openpyxl creates
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Summary sheet (index 0) ───────────────────────────────────────────
    _write_summary_sheet(
        wb, results, platform,
        har_filename=os.path.basename(har_path),
        cg=cg, gm=gm, cl=cl,
    )

    # ── One sheet per section, in canonical order ─────────────────────────
    for section in SECTION_ORDER:
        rows = results.get(section, [])
        if not rows:
            continue
        _write_section_sheet(wb, section, rows, debug=debug)

    # ── Any non-standard sections appended at the end ─────────────────────
    for section, rows in results.items():
        if section in SECTION_ORDER or not rows:
            continue
        _write_section_sheet(wb, section, rows, debug=debug)

    wb.save(out_path)
    return out_path
